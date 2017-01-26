#  Will copy serial nbr from excel and alt+tab to
#  chrome and paste. Enter. Tab twice, if there is a "inactive"
#  on the webpage. Else, Alert me.
import openpyxl
import os
import ctypes
import time


class AutoJob:
    def __init__(self):
        full_or_not = raw_input("Type Full or Sheet Name: ")
        if full_or_not != 'Full':
            start_here = raw_input("What cell do you want to start at? (Integer): ")
            try:
                if type(int(start_here)) != type(1):
                    self.get_sheet(full_or_not, 1)
                else:
                    self.get_sheet(full_or_not, int(start_here))
            except ValueError:
                self.get_sheet(full_or_not, 1)
        else:
            self.get_file()

    def alt_tab(self):
        ctypes.windll.user32.keybd_event(0x12, 0, 0, 0)  # alt
        ctypes.windll.user32.keybd_event(0x09, 0, 0, 0)  # tab
        time.sleep(1)  # wait 2
        ctypes.windll.user32.keybd_event(0x09, 0, 2, 0)  # !alt
        ctypes.windll.user32.keybd_event(0x012, 0, 2, 0)  # !tab

    def tab(self):
        ctypes.windll.user32.keybd_event(0x09, 0, 0, 0)  # tab
        ctypes.windll.user32.keybd_event(0x09, 0, 2, 0)

    def paste(self):
        ctypes.windll.user32.keybd_event(0xA2, 0, 0, 0)  # ctrl
        ctypes.windll.user32.keybd_event(0x56, 0, 0, 0)  # v
        ctypes.windll.user32.keybd_event(0xA2, 0, 2, 0)  # !ctrl
        ctypes.windll.user32.keybd_event(0x56, 0, 2, 0)  # !v

    def enter(self):
        ctypes.windll.user32.keybd_event(0x0D, 0, 0, 0)  # enter
        ctypes.windll.user32.keybd_event(0x0D, 0, 2, 0)

    def first_click(self):
        for i in range(2):
            self.tab()
        self.enter()

    def change_status(self):
        for i in range(3):
            time.sleep(0.5)
            print("Status")
            self.tab()
            time.sleep(0.2)
        ctypes.windll.user32.keybd_event(0x55, 0, 0, 0)  # u
        ctypes.windll.user32.keybd_event(0x55, 0, 2, 0)

    def change_storage(self):
        for i in range(6):
            print("Storage Location")
            self.tab()
            time.sleep(0.2)
        ctypes.windll.user32.keybd_event(0x53, 0, 0, 0)  # s
        ctypes.windll.user32.keybd_event(0x53, 0, 2, 0)
        ctypes.windll.user32.keybd_event(0x45, 0, 0, 0)  # e
        ctypes.windll.user32.keybd_event(0x45, 0, 2, 0)
        ctypes.windll.user32.keybd_event(0x4E, 0, 0, 0)  # n
        ctypes.windll.user32.keybd_event(0x4E, 0, 2, 0)
        ctypes.windll.user32.keybd_event(0x54, 0, 0, 0)  # t
        ctypes.windll.user32.keybd_event(0x54, 0, 2, 0)
        ctypes.windll.user32.keybd_event(0x20, 0, 0, 0)  # space
        ctypes.windll.user32.keybd_event(0x20, 0, 2, 0)
        ctypes.windll.user32.keybd_event(0x46, 0, 0, 0)  # f
        ctypes.windll.user32.keybd_event(0x46, 0, 2, 0)

    def change_date(self):
        for i in range(12):
            print("End Cust Maint")
            self.tab()
            time.sleep(0.2)
        ctypes.windll.user32.keybd_event(0x31, 0, 0, 0)  # 1
        ctypes.windll.user32.keybd_event(0x31, 0, 2, 0)
        time.sleep(0.2)
        ctypes.windll.user32.keybd_event(0x28, 0, 0, 0)  # Down arrow
        ctypes.windll.user32.keybd_event(0x28, 0, 2, 0)
        time.sleep(0.2)
        self.enter()

    def save_exit(self):
        for i in range(6):
            print('Save')
            self.tab()
            time.sleep(0.2)
        self.enter()

    def cancel_butt(self):
        for i in range(8):
            print('Cancel')
            self.tab()
            time.sleep(0.2)
        self.enter()
        time.sleep(0.5)
        self.enter()

    # alt tabs through and checks the excel
    def check_tab(self):
        ctypes.windll.user32.keybd_event(0x12, 0, 0, 0)  # alt
        ctypes.windll.user32.keybd_event(0x09, 0, 0, 0)  # tab
        ctypes.windll.user32.keybd_event(0x09, 0, 0, 0)  # tab
        time.sleep(2)
        ctypes.windll.user32.keybd_event(0x09, 0, 2, 0)  # !alt
        ctypes.windll.user32.keybd_event(0x012, 0, 2, 0)  # !tab
        time.sleep(1)
        ctypes.windll.user32.keybd_event(0xA2, 0, 0, 0)  # ctrl
        ctypes.windll.user32.keybd_event(0x22, 0, 0, 0)  # pgup
        ctypes.windll.user32.keybd_event(0xA2, 0, 2, 0)  # ctrl
        ctypes.windll.user32.keybd_event(0x22, 0, 2, 0)  # pgup
        time.sleep(0.5)
        ctypes.windll.user32.keybd_event(0x12, 0, 0, 0)  # alt
        ctypes.windll.user32.keybd_event(0x09, 0, 0, 0)  # tab
        time.sleep(1)  # wait 2
        ctypes.windll.user32.keybd_event(0x09, 0, 2, 0)  # !alt
        ctypes.windll.user32.keybd_event(0x012, 0, 2, 0)  # !tab

    # function to copy a variable to the clipboard
    def add_to_clip(self, text):
        if type(text) == str:
            command = 'echo ' + text + '| clip'
            os.system(command)

    # This does all the leg work
    def human_job(self, serial):
        self.add_to_clip(serial)
        self.paste()
        time.sleep(1)
        self.enter()
        time.sleep(1)
        self.first_click()
        time.sleep(1)
        self.change_status()
        time.sleep(1)
        self.change_storage()
        time.sleep(1)
        self.change_date()
        time.sleep(1)
        self.save_exit()
        time.sleep(3)

    def get_file(self):
        self.alt_tab()
        wb = openpyxl.load_workbook('test.xlsx')
        # print(type(wb))
        # fetch the sheets of the workbook
        for x in wb.get_sheet_names():
            curr_sheet = wb.get_sheet_by_name(x)
            for i in range(1, curr_sheet.max_row+1):
                if type(curr_sheet.cell(row=i, column=1).value) == unicode:
                    reg_string = curr_sheet.cell(row=i, column=1).value.encode('ascii', 'ignore')
                    self.human_job(reg_string)
                    # print(currSheet.cell(row=i, column=1).value)
            time.sleep(1)
            self.check_tab()

    def get_sheet(self, starting_sheet, start_here):
        self.alt_tab()
        wb = openpyxl.load_workbook('second.xlsx')
        sheet_index = wb.get_sheet_names().index(starting_sheet)
        for x in wb.get_sheet_names()[sheet_index:len(wb.get_sheet_names())+1]:
            curr_sheet = wb.get_sheet_by_name(x)
            for i in range(start_here, curr_sheet.max_row + 1):
                if type(curr_sheet.cell(row=i, column=1).value) == unicode:
                    reg_string = curr_sheet.cell(row=i, column=1).value.encode('ascii', 'ignore')
                    self.human_job(reg_string)
                    # print(currSheet.cell(row=i, column=1).value)
            time.sleep(1)
            self.check_tab()

AutoJob()
